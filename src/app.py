import openpyxl
import csv
from tabulate import tabulate

# colocando el nombre del archivo leer
comites = ["SPECPOL", "UNODC", "CRISIS", "COMISION"]
excel_dataFrame = openpyxl.load_workbook("Planilla.xlsx")
bigData = []
class delegado:
    nombre: str
    phone: int
    def __init__(self,nombre, phone):
        self.nombre = nombre
        self.phone = phone
        
def ObtenerNumeros(data_frame):
    data = []
    for ro in range(2, data_frame.max_row):
        valorColumna = (data_frame.cell(row=ro, column=5).value)
        valorName = data_frame.cell(row=ro, column=3).value
        condicion = str(valorColumna).lower() != "none" and str(valorName).lower() != "none"
        if(condicion):
            valorColumna = "+58" + str(int(valorColumna)) 
            valorName = valorName + " " + data_frame.title
            data.append(delegado(nombre=valorName, phone=valorColumna))
    return data

arr = []
arch = open('nums.md', 'w')
arch.write("hola mundo")
for d in range(0, len(comites)):
    comite = comites[d]
    print(comite)
    data2 = ObtenerNumeros(excel_dataFrame[comite])
    
    bigData.append(data2)
for d in range(0, len(bigData)):
    for j in range(0, len(bigData[d])):
        arr.append([bigData[d][j].nombre, bigData[d][j].phone])
        arch.write(bigData[d][j].phone + "\n")

arch.close()
